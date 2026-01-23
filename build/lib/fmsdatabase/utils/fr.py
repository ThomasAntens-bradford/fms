# --- Standard Library ---
import io
import os
import re
import time
import traceback
from datetime import datetime
import json

# --- Third-Party Libraries ---
import numpy as np
import openpyxl
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

# --- Script Path Handling ---
if __name__ == "__main__":
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# --- Local Imports ---
from ..db import (
    FRCertification,
    AnodeFR,
    CathodeFR,
    ManifoldStatus,
    LPTCalibration,
    TestingTools
)

from .ocr_reader import OCRReader
from .textract import TextractReader
from .general_utils import (
    FRParts,
    ManifoldProgressStatus,
    extract_total_amount,
    FRStatus
)

# --- Type Checking ---
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from sqlalchemy.orm import Session
    from ..fms_data_structure import FMSDataStructure

class FRListener(FileSystemEventHandler):
    """
    Base class for listening to new FR certification files in a specified directory.
    Inherits from FileSystemEventHandler to handle file system events.

    Attributes:
        path (str): The directory path to monitor for new files.
        observer (Observer): Watchdog observer to monitor file system events.
        parts (bool): Flag indicating if the processed file contains parts data.
    """

    def __init__(self, path: str):
        self.path = path
        self.observer = Observer()
        self.observer.schedule(self, path=path, recursive=True)
        self.observer.start()
        self.parts = False

    def on_created(self, event) -> None:
        """
        Event handler for file creation events.
        Distinguishes between different types of FR certification files and processes them accordingly.
        """
        if not event.is_directory and event.src_path.endswith('.pdf'):
            try:
                # Process the data and save to database
                if self.path.endswith('certifications'):
                    
                    if not event.src_path.endswith('.pdf'):
                        return
                    print(f"New FR certification file detected: {event.src_path}")
                    self.pdf_file = event.src_path
                    companies = ['veld laser', 'branl', 'veldlaser']
                    if any(company in self.pdf_file.lower() for company in companies):
                        self.fr_data = FRData(pdf_file=self.pdf_file)
                        self.fr_data.get_ocr_certification()
                        self.processed = True
                    elif 'ejay filtration' in self.pdf_file.lower():
                        self.parts = True
                        self.fr_data = FRData(pdf_file = self.pdf_file)
                        self.fr_data.get_filter_certification()
                        self.processed = True
                    elif 'sk' in self.pdf_file.lower() or 'sk technology' in self.pdf_file.lower():
                        self.parts = True
                        self.fr_data = FRData(pdf_file = self.pdf_file)
                        self.fr_data.get_outlet_certification()
                        self.processed = True
                    
            except Exception as e:
                print(f"Error processing file {event.src_path}: {e}")
                traceback.print_exc()

class FRData:
    """
    Base class for handling flow restrictor (FR) certification data extraction and processing.
    Extracts relevant information from PDF certification files.
    Extracts and standardizes existing data (from tests/assembly) from Excel files.

    Attributes:
        pdf_file (str): 
            Path to the PDF certification file.
        anode_reference_orifice (float): 
            Reference orifice diameter for anode FRs.
        cathode_reference_orifice (float): 
            Reference orifice diameter for cathode FRs.
        reference_thickness (float): 
            Reference thickness for FRs.
        thickness_tolerance (float): 
            Allowed thickness tolerance for FRs.
        cathode_excel (str): 
            Path to the cathode FR Excel data file.
        anode_excel (str):
            Path to the anode FR Excel data file.
        anode_drawing (str):
            Standard drawing number for anode FRs.
        cathode_drawing (str):
            Standard drawing number for cathode FRs.
        extracted_fr_parts (dict):
            Extracted FR parts data from the certification.
        fr_test_results (list):
            Extracted FR test results from Excel files.
        company (str):
            Company name associated with the certification.
        filter_total_lines (str):
            OCR-extracted text lines for filter certifications.
        outlet_total_lines (str):
            OCR-extracted text lines for outlet certifications.
        total_lines (list):
            OCR-extracted text lines for restrictor certifications.

    Methods:
        extract_fr_parts():
            Extracts FR parts data from OCR text lines. ***REMARK: USES OCRReader CLASS, NOT TEXTRACT!***
        get_ocr_certification():
            Extracts certification data from the PDF file using OCR. ***REMARK: USES OCRReader CLASS, NOT TEXTRACT!***
        extract_filter(part):
            Extracts filter part data from Textract results.
        extract_outlet(part):
            Extracts outlet part data from Textract results.
        extract_restrictor(part):
            Extracts restrictor part data from Textract results.
        get_certification(total_lines):
            Determines the certification type and extracts relevant data, feeds to part-specific methods.
        extract_filter_quantity():
            Extracts filter quantity from OCR text lines. ***REMARK: USES OCRReader CLASS, NOT TEXTRACT!***
        extract_outlet_quantity():
            Extracts outlet quantity from OCR text lines. ***REMARK: USES OCRReader CLASS, NOT TEXTRACT!***
        get_filter_certification():
            Processes filter certification PDF files. ***REMARK: USES OCRReader CLASS, NOT TEXTRACT!***
        get_outlet_certification():
            Processes outlet certification PDF files. ***REMARK: USES OCRReader CLASS, NOT TEXTRACT!***
        excel_data_loop(row, type):
            Processes a row of Excel data for either anode or cathode FRs.
        extract_data_from_excel():
            Extracts FR test results from Excel files.
    """

    def __init__(self, pdf_file: str = None, anode_reference_orifice: float = 0.07095, cathode_reference_orifice: float = 0.01968, 
                 reference_thickness: float = 0.25, thickness_tolerance: float = 0.01, cathode_excel: str = "Excel_templates/cathode_copy.xlsx", 
                 anode_excel: str = "Excel_templates/anode_copy.xlsx"):
        
        self.pdf_file = pdf_file
        self.certification = None
        self.anode_fr = None
        self.cathode_fr = None
        self.anode_reference_orifice = anode_reference_orifice
        self.cathode_reference_orifice = cathode_reference_orifice
        self.reference_thickness = reference_thickness
        self.thickness_tolerance = thickness_tolerance
        self.allowed_thickness_deviation = self.thickness_tolerance/reference_thickness*100
        self.anode_drawing = '20025.10.18-R4-005'
        self.drawing_reference = None
        self.cathode_drawing = '20025.10.18-R4-001'
        self.extracted_fr_parts = {}
        self.fr_test_results = {}
        self.anode_excel = anode_excel
        self.company = None
        self.cathode_excel = cathode_excel
        self.filter_total_lines = ''
        self.outlet_total_lines = ''
        self.total_lines = []

    def extract_fr_parts(self) -> dict:
        """
        Extracts FR parts data from OCR text lines.
        Returns:
            dict: Extracted FR parts data with serial numbers as keys.
        """
        fr_parts = {}

        serial_re = re.compile(r'^\d{2,3}$')
        number_re = re.compile(r'^[\d\s.,]{3,10}$')

        # Pre-clean lines: remove empty and obviously invalid lines
        lines = [line.strip() for line in self.lines.split('\n') if line.strip()]
        seen_serials = set()

        if self.drawing:
            if self.drawing.endswith('005'):
                part_type = "anode"
                reference_orifice = self.anode_reference_orifice
            elif self.drawing.endswith('001'):
                part_type = "cathode"
                reference_orifice = self.cathode_reference_orifice
            drawing = self.drawing

        def parse_number(raw: str) -> float:
            """Parses and corrects OCR-mangled floats like '0,242' or '0.0204'."""
            raw = raw.replace(' ', '').replace(',', '.')
            try:
                val = float(raw)
                if val > 1.0:  # Ignore unlikely OCR results
                    return None
                return round(val, 5)
            except ValueError:
                return None

        i = 0
        while i < len(lines) - 2:
            serial = lines[i]
            val1_raw = lines[i + 1]
            val2_raw = lines[i + 2]

            if serial_re.match(serial) and serial not in seen_serials:
                val1 = parse_number(val1_raw)
                val2 = parse_number(val2_raw)

                if val1 is not None and val2 is not None:
                    seen_serials.add(serial)

                    if val2 > 0.4:
                        val2 /= 10
                    # Heuristic: thickness is usually slightly more than orifice
                    thickness, orifice = sorted([val1, val2], reverse=True)

                    # Identify part type
                    if abs(orifice - 0.070) < 0.01:
                        part_type = "anode"
                        drawing = self.anode_drawing if not self.drawing else self.drawing
                        reference_orifice = self.anode_reference_orifice
                    elif abs(orifice - 0.020) < 0.01:
                        part_type = "cathode"
                        drawing = self.cathode_drawing if not self.drawing else self.drawing
                        reference_orifice = self.cathode_reference_orifice
                    else:
                        part_type = "unknown"
                        drawing = self.cathode_drawing if not self.drawing else self.drawing
                        reference_orifice = self.cathode_reference_orifice

                    deviation = (orifice - reference_orifice) / reference_orifice * 100
                    deviation_thickness = thickness < self.reference_thickness

                    # Status determination
                    if abs(deviation) > 10 and deviation_thickness:
                        status = FRStatus.DIFF_GEOMETRY
                    elif abs(deviation) > 10:
                        status = FRStatus.DIFF_ORIFICE
                    elif deviation_thickness:
                        status = FRStatus.DIFF_THICKNESS
                    else:
                        status = FRStatus.OK

                    fr_parts[f'{self.certification}-{str(serial).zfill(3)}'] = {
                        "thickness": thickness,
                        "orifice": orifice,
                        "deviation": deviation,
                        "status_geometry": status,
                        "fr_type": part_type,
                        "drawing": drawing
                    }

                    i += 3
                    continue

            i += 1 

        return fr_parts
    
    def get_ocr_certification(self) -> None:
        ocr_reader = OCRReader(pdf_file=self.pdf_file)
        self.lines, self.drawing = ocr_reader.read_scanned_page_veldlaser()
        self.certification = ocr_reader.certification
        if self.lines and self.drawing:
            self.extracted_fr_parts = self.extract_fr_parts()
        # print(self.extracted_fr_parts)
        # print(len(self.extracted_fr_parts))

    def extract_filter(self, part: str) -> dict:
        """
        Extracts filter part data from Textract results.
        """
        default_drawing = '84900P-395-05-316L'
        results = {
            part:
            {
                'amount': self.total_amount,
                'certification': self.certification,
                'drawing': self.drawing_reference if self.drawing_reference else default_drawing
            }
        }
        return results
    
    def extract_outlet(self, part: str) -> dict:
        """
        Extracts outlet part data from Textract results.
        """
        default_drawing = '20025.10.21-R5'
        results = {
            part:
            {
                'amount': self.total_amount,
                'certification': self.certification,
                'drawing': self.drawing_reference if self.drawing_reference else default_drawing
            }
        }
        return results
    
    def extract_with_thickness(self) -> dict:
        """
        Extracts restrictor part data from Textract results with thickness measurements.
        """
        results = {}
        seen_serials = set()
        sns = [idx for idx, i in enumerate(self.total_lines) if i == 'sn' or (i == 'part' and self.total_lines[idx+1] == 'm1')]
        first_sn = sns[0] if sns else 0
        last_sn = sns[-1] if sns else len(self.total_lines)
        serial_match = re.compile(r'^\d{2,3}$')
        for idx, line in enumerate(self.total_lines):
            if serial_match.match(line):
                if line not in seen_serials and (first_sn < idx < last_sn + 6):
                    seen_serials.add(line)
                    val1 = self.total_lines[idx + 1].strip().replace(',', '.')
                    val2 = self.total_lines[idx + 2].strip().replace(',', '.')
                    try:
                        thickness = float(val1)
                        orifice = float(val2)
                    except ValueError:
                        continue

                    if thickness > 0.4:
                        thickness /= 10
                    
                    if orifice > 0.2:
                        orifice /= 10

                    results[f"{self.certification}-{str(line).zfill(3)}"] = {
                        'thickness': thickness,
                        'orifice': orifice
                    }
        return results
    
    def extract_restrictor_measurements(self) -> dict:
        """
        Extracts restrictor part data from Textract results without thickness measurements.
        Only parses consecutive serial number (integer) and measurement pairs starting after
        'results of measurement:' and after skipping repeated 'sn', '1' headers.
        """
        results = {}
        parsing = False
        i = 0
        header_count = 0

        while i < len(self.total_lines):
            if not parsing and self.total_lines[i].lower() == "results of measurement:":
                parsing = True
                i += 1
                continue

            if parsing:
                if self.total_lines[i].lower() == 'sn' and i + 1 < len(self.total_lines) and self.total_lines[i + 1].lower() == '1':
                    header_count += 1
                    i += 2
                    continue

                if header_count >= 4:
                    if self.total_lines[i].isdigit():  
                        sn = int(self.total_lines[i])
                        if i + 1 < len(self.total_lines):
                            try:
                                meas = float(self.total_lines[i + 1])
                                results[f"{self.certification}-{str(sn).zfill(3)}"] = {
                                    'orifice': meas
                                }
                                i += 2
                                continue
                            except ValueError:
                                break
                    else:
                        break
            i += 1

        return results

    def extract_restrictor(self, part: str) -> dict:
        """
        Extracts restrictor part data from Textract results.
        Determines whether the part is an anode or cathode based on orifice size, 
        and assigns the corresponding serial numbers.
        """
        if any(i == 'm2' for i in self.total_lines):
            results_dict = self.extract_with_thickness()
        else:
            results_dict = self.extract_restrictor_measurements()

        if self.drawing_reference:
            if self.drawing_reference.endswith('005'):
                part_type = "anode"
                reference_orifice = self.anode_reference_orifice
            elif self.drawing_reference.endswith('001'):
                part_type = "cathode"
                reference_orifice = self.cathode_reference_orifice

        for key, value in results_dict.items():
            thickness = value.get('thickness', None)
            orifice = value.get('orifice', None)

            if abs(orifice - self.anode_reference_orifice) < 0.01:
                part_type = "anode"
                self.drawing_reference = self.anode_drawing if not self.drawing_reference else self.drawing_reference
                reference_orifice = self.anode_reference_orifice
            elif abs(orifice - self.cathode_reference_orifice) < 0.01:
                part_type = "cathode"
                self.drawing_reference = self.cathode_drawing if not self.drawing_reference else self.drawing_reference
                reference_orifice = self.cathode_reference_orifice
            else:
                part_type = "unknown"
                self.drawing_reference = self.cathode_drawing if not self.drawing_reference else self.drawing_reference
                reference_orifice = self.cathode_reference_orifice

            deviation = (orifice - reference_orifice) / reference_orifice * 100
            if thickness is not None:
                deviation_thickness = np.abs(thickness - self.reference_thickness) / self.reference_thickness * 100
            else:
                deviation_thickness = 0

            # Status determination
            if abs(deviation) > 10 and bool(thickness) and deviation_thickness > self.allowed_thickness_deviation:
                status = FRStatus.DIFF_GEOMETRY
            elif abs(deviation) > 10:
                status = FRStatus.DIFF_ORIFICE
            elif bool(thickness) and deviation_thickness > self.allowed_thickness_deviation:
                status = FRStatus.DIFF_THICKNESS
            else:
                status = FRStatus.OK

            results_dict[key] = {
                'thickness': thickness,
                'orifice': orifice,
                'deviation': deviation,
                'status_geometry': status,
                'fr_type': part_type,
                'drawing': self.drawing_reference
            }

        return results_dict

    def get_certification(self, total_lines: list[str]) -> None:
        """
        Determines the certification type and extracts relevant data, feeds to part-specific methods.
        Args:
            total_lines (list[str]): Extracted text lines from the certification PDF, using Textract.
        """
        part = None
        self.function_map = {
            FRParts.FILTER.value: self.extract_filter,
            FRParts.OUTLET.value: self.extract_outlet,
            FRParts.RESTRICTOR.value: self.extract_restrictor
        }
        self.total_lines = total_lines
        print(total_lines)
        match = re.search(r'C\d{2}-\d{4}', os.path.basename(self.pdf_file))
        self.certification = match.group(0) if match else None
        for line in self.total_lines:
            if self.company == 'ejay filtration':
                drawing_match = re.search(r'\b[0-9]{5}[A-Z]?-?[0-9]{3}-[0-9]{2}-[A-Z0-9]+', line)
            elif self.company == 'branl':
                drawing_match = re.search(
                    r'([0-9]{5}\.[0-9]{2}\.[a-zA-Z0-9.]{2,3}-R[0-9]+-\d{3})(?:\s+\w+)?',
                    line,
                    re.IGNORECASE
                )
            else:
                drawing_match = re.search(r'([0-9]{5}\.[0-9]{2}\.[a-zA-Z0-9]{2,3}-R[0-9]+)(?:\s+\w+)?', line, re.IGNORECASE)

            if drawing_match:
                if not self.drawing_reference:
                    self.drawing_reference = drawing_match.group(0).upper()


            if any(i for i in self.function_map if i in line):
                if not part:
                    part = next(i for i in self.function_map if i in line)

        try:
            #TODO if part is restrictor outlet, try the certificate of conformity
            # Find index of element that contains 'totaal aantal'
            totaal_index = next(
                idx for idx, val in enumerate(self.total_lines)
                if re.search(r'totaal\s*aantal', val, re.IGNORECASE)
            )

            # Look ahead to find the next item with digits
            for next_val in self.total_lines[totaal_index+1 : totaal_index+5]:
                if re.search(r'\d', next_val):  # contains a digit
                    self.total_amount = int(float(next_val.strip().replace(',', '.')))
                    break
            
        except (IndexError, ValueError, StopIteration):
            try:
                quantity_pattern = r"quantity(?: supplied)?:\s*(\d+)"
                quantity_line = next((i for i in self.total_lines if re.search(quantity_pattern, i, re.IGNORECASE)), None)
                if quantity_line:
                    match = re.search(quantity_pattern, quantity_line, re.IGNORECASE)
                    if match:
                        self.total_amount = int(match.group(1))
            except (IndexError, ValueError):
                quantity_index = self.total_lines.index('quantity shipped:')
                self.total_amount = int(self.total_lines[quantity_index + 1].strip().replace(',', '').replace('.', ''))

        self.extracted_fr_parts = self.function_map[part](part)

    def extract_filter_quantity(self) -> dict:
        """
        Extracts filter quantity and drawing number from OCR text lines.
        """
        text = self.filter_total_lines
        part_name = 'ejay filter'
        default_drawing = '84900P-395-05-316L'

        text_lower = text.lower()
        result = {}
        drawing_found = default_drawing
        amount_found = None
        part_found = None

        # Look for drawing number near "ejay filter"
        for line in text.splitlines():
            if part_name in line.lower():
                part_found = True
                match = re.search(r'\b[0-9]{5}[A-Z]?-?[0-9]{3}-[0-9]{2}-[A-Z0-9]+', line)
                if match:
                    drawing_found = match.group(0)
                break  # Only first match needed
        
        if not part_found:
            return {}
        
        # Look for "totaal aantal" to extract amount
        total_match = re.search(r'totaal aantal\s*[:=]*\s*(\d{1,3}(?:[.,]\d{3})*|\d+)', text_lower)
        if total_match:
            try:
                total_str = total_match.group(1).replace('.', '').replace(',', '.')
                amount_found = int(round(float(total_str)))
            except:
                amount_found = None

        if amount_found is not None and part_found:
            result[part_name] = {
                'amount': amount_found,
                'drawing': drawing_found
            }

        return result
    
    def extract_outlet_quantity(self) -> dict:
        """
        Extracts outlet quantity and drawing number from OCR text lines.
        """
        text = self.outlet_total_lines
        part_name = 'restrictor outlet'
        default_drawing = '20025.10.21-R5'
        text_lower = text.lower()
        result = {}
        drawing_found = default_drawing
        amount_found = None
        part_found = None
        # Look for drawing number near "restrictor outlet"
        for line in text.splitlines():
            if part_name in line.lower():
                match = re.search(r'([0-9]{5}\.[0-9]{2}\.[0-9]{2,3}-[A-Z0-9]+)', line)
                if match:
                    drawing_found = match.group(1)

                part_found = True
                break  # Only first match needed

        # Look for "totaal aantal" to extract amount

        amount_found = extract_total_amount(text_lower)
        if amount_found is not None and part_found:
            result[part_name] = {
                'amount': amount_found,
                'drawing': drawing_found
            }

        return result
    
    def get_filter_certification(self) -> None:
        """
        Processes filter certification PDF files.
        """
        ocr_reader = OCRReader(pdf_file=self.pdf_file)
        ocr_reader.main_delivery_slip_reader('filter')
        self.filter_total_lines = ocr_reader.total_lines
        self.certification = ocr_reader.certification
        if self.filter_total_lines:
            self.extracted_fr_parts = self.extract_filter_quantity()

    def get_outlet_certification(self) -> None:
        """
        Processes outlet certification PDF files.
        """
        ocr_reader = OCRReader(pdf_file=self.pdf_file)
        ocr_reader.main_delivery_slip_reader('outlet')
        self.outlet_total_lines = ocr_reader.total_lines
        self.certification = ocr_reader.certification
        if self.outlet_total_lines:
            self.extracted_fr_parts = self.extract_outlet_quantity()

    def get_tools(self, tools_path: str, type: str, trs_reference: str):
        """
        Takes trs reference tool data and assigns them to the current FR being processed.

        :param type: FR type
        :type type: str
        :param trs_reference: Reference to the TRS that the FR is included in.
        :type trs_reference: str

        :Returns:
            components type:
        """
        import random
        trs_tools = {}
        if os.path.exists(tools_path):
            with open(tools_path, 'r', encoding='utf-8') as f:
                trs_tools = json.load(f)

        relevant_tools: list[dict[str, str]] = trs_tools.get(trs_reference, [])
        if not relevant_tools:
            return []
        
        components = []
        
        possible_TR = [i for i in relevant_tools if i["description"] == "Temperature Recorder"]
        chosen_TR = random.choice(possible_TR)
        components.append({"description": "_".join(chosen_TR.get("description").split()).lower(), "serial_number": chosen_TR.get("serial_number"), "model": chosen_TR.get("model")})

        possible_FS = [i for i in relevant_tools if i["description"] == "Mass Flow Sensor" and i["type"] == type]
        chosen_FS = random.choice(possible_FS)
        components.append({"description": "_".join(chosen_FS.get("description").split()).lower(), "serial_number": chosen_FS.get("serial_number"), "model": chosen_FS.get("model")})

        possible_IP = [i for i in relevant_tools if i["description"] == "Inlet Pressure Controller"]
        chosen_IP = random.choice(possible_IP)
        components.append({"description": "_".join(chosen_IP.get("description").split()).lower(), "serial_number": chosen_IP.get("serial_number"), "model": chosen_IP.get("model")})

        possible_OP = [i for i in relevant_tools if i["description"] == "Outlet Pressure Controller"]
        chosen_OP = random.choice(possible_OP)
        components.append({"description": "_".join(chosen_OP.get("description").split()).lower(), "serial_number": chosen_OP.get("serial_number"), "model": chosen_OP.get("model")})
        return components

    def excel_data_loop(self, row: tuple, type: str, tools_path: str) -> dict:
        """
        Processes a row of Excel data for either anode or cathode FRs.
        """
        if type == 'anode':
            cert, drawing, id, thickness, orifice, deviation, radius, temperature, allocated, remark = row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[14], row[15]
            flow_rates = [row[9], row[10], row[11], row[12]]
            trs_reference = row[13]
            if all(i is None for i in flow_rates):
                flow_rates = []

        elif type == 'cathode':
            if len(row) > 17:
                end_index = 16
            else:
                end_index = 15
            cert, drawing, id, thickness, orifice, deviation, radius, temperature, allocated, remark = row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[end_index -7], row[end_index-1], row[end_index]
            flow_rates = [row[end_index - 6], row[end_index - 5], row[end_index - 4], row[end_index - 3]]
            trs_reference = row[end_index - 2]
            if all(i is None for i in flow_rates):
                flow_rates = []

        if flow_rates and not cert and not id and self.fr_test_results:
            if "flow_rates" in self.fr_test_results[-1]:
                self.fr_test_results[-1]["flow_rates"] = flow_rates
            return {}  

        pressures = [1, 1.5, 2, 2.4]
        set_id = None
        if 'set' in str(allocated).lower():
            match = re.search(r"set[-\s]*(\d+)", allocated, re.IGNORECASE)
            if match:
                set_id = match.group(1)

        if deviation and abs(deviation) > 5 and thickness and thickness < self.reference_thickness:
            status_geometry = FRStatus.DIFF_GEOMETRY
        elif deviation and abs(deviation) > 5:
            status_geometry = FRStatus.DIFF_ORIFICE
        elif thickness and thickness < self.reference_thickness:
            status_geometry = FRStatus.DIFF_THICKNESS
        else:
            status_geometry = FRStatus.OK

        if not cert or not id:
            return {}
        
        fr_id = f'{cert}-{str(id).zfill(3)}'
        if fr_id in self.previous_ids:
            return {}
                
        if any(re.match(r'^(n/?a)$', str(val).strip(), re.IGNORECASE) for val in flow_rates):
            return {}
        
        tools = self.get_tools(tools_path, type, trs_reference)

        fr_dict = {
            'fr': type.capitalize(),
            'serial_number': fr_id,
            'thickness': thickness,
            'orifice_diameter': orifice,
            'gas_type': 'Xe',
            'deviation': deviation,
            'radius': radius,
            'temperature': temperature,
            'pressures': pressures,
            'flow_rates': flow_rates,
            'status_geometry': status_geometry,
            'date': None,
            'remark': remark,
            'allocated': allocated,
            'set_id': set_id,
            'trs_reference': trs_reference,
            'tools': tools,
            'drawing': drawing
        }

        return fr_dict

    def extract_data_from_excel(self, operator: str = "JKR", tools_path: str = "trs_tools_fr") -> list:
        """
        Extracts FR test results from Excel files.
        Returns:
            list: List of extracted FR test results.
        """
        self.fr_test_results = []
        self.previous_ids = []
        operator = operator

        wb_anode = openpyxl.load_workbook(self.anode_excel, data_only=True)
        if 'Anode FR Database' in wb_anode.sheetnames:
            wb_anode.active = wb_anode['Anode FR Database']
        else:
            wb_anode.active = wb_anode.active

        ws = wb_anode.active
        for idx, row in enumerate(wb_anode.active.iter_rows(min_row=3, values_only=True)):
            if all(cell is None for cell in row):
                break
            fr_dict = self.excel_data_loop(row, type='anode', tools_path=tools_path)
            cell = ws.cell(row=idx+1, column=1)
            filled = cell.fill.fill_type == 'solid'
            if filled and bool(fr_dict.get("flow_rates", [])):
                operator = "NRN"
            elif bool(fr_dict.get("flow_rates", [])):
                operator = "JKR"
            else:
                operator = ""

            if fr_dict:  
                fr_dict["operator"] = operator
                self.fr_test_results.append(fr_dict)

        wb_cathode = openpyxl.load_workbook(self.cathode_excel, data_only=True)
        if 'Cathode FR Database' in wb_cathode.sheetnames:
            wb_cathode.active = wb_cathode['Cathode FR Database']
        else:
            wb_cathode.active = wb_cathode.active

        ws = wb_cathode.active
        for idx, row in enumerate(wb_cathode.active.iter_rows(min_row=3, values_only=True)):
            if all(cell is None for cell in row):
                break
            fr_dict = self.excel_data_loop(row, type='cathode', tools_path=tools_path)
            cell = ws.cell(row=idx+1, column=1)
            filled = cell.fill.fill_type == 'solid'
            if filled and bool(fr_dict.get("flow_rates", [])):
                operator = "NRN"
            elif bool(fr_dict.get("flow_rates", [])):
                operator = "JKR"
            else:
                operator = ""

            if fr_dict:  
                fr_dict["operator"] = operator
                self.fr_test_results.append(fr_dict)
            if fr_dict: 
                self.fr_test_results.append(fr_dict)

        return self.fr_test_results

class FRLogicSQL:
    """
    Class for handling SQL logic related to flow restrictor (FR) certifications and test results.
    Handles all the database interactions for the FRs.

    Attributes
    ------------
        Session: 
            SQLAlchemy session object.
        fr_certifications (str): 
            Table name for FR certifications.
        anode_target_15 (float): 
            Target flow rate for anode FRs at 1.5 bar.
        anode_target_24 (float): 
            Target flow rate for anode FRs at 2.4 bar.
        cathode_target_15 (float): 
            Target flow rate for cathode FRs at 1.5 bar.
        cathode_target_24 (float): 
            Target flow rate for cathode FRs at 2.4 bar.
        min_radius (float): 
            Minimum acceptable radius for FRs.
        max_radius (float): 
            Maximum acceptable radius for FRs.
        fr_id: 
            Current FR ID being processed.
        fr_test_results (dict): 
            Dictionary to hold FR test results.
        fr_matching_dict (dict): 
            Dictionary to hold matching information for FRs.
        fr_matching_ratios (dict): 
            Dictionary to hold matching ratios for FRs.
        lines (str): 
            OCR-extracted text lines from certification PDFs.
        manifold_assembly_data (dict): 
            Dictionary to hold manifold assembly data.
        xenon_density (float): 
            Density of xenon gas at standard conditions.
        krypton_density (float): 
            Density of krypton gas at standard conditions.
    
    Methods
    ----------
        listen_to_fr_certifications():
            Starts listening for new FR certification files.
        update_fr_certification(fr_data: FR_data = None):
            Updates the database with new FR certification data.
        get_flow_restrictors():
            Retrieves all unallocated flow restrictors from the database.
        get_status(flow_rates, type): 
            Determines the status of FRs based on flow rates and type.
        get_radius_status(status, radius):
            Adjusts the FR status based on the radius size and the current status.
        flow_test_inputs(): 
            Displays the UI for the FR testing inputs.
        plot_fr_results(pressures, flow_rates, type, gas_type, temperature):
            Plots the FR test results (flow rate vs pressure).
        hagen_poiseuille(gas_type, flow_rate, thickness, orifice_diameter, viscosity):
            Calculates the theoretical pressure drop based on Hagen-Poiseuille flow.
        fr_test_result_sql(test_results, excel_extraction, session):
            Helper function to handle the database insertion of FR test results.
        update_fr_test_results(excel_extraction, anode_path, cathode_path):
            Updates the database with new FR test results from Excel files if specified,
            otherwise uses results from the flow test input UI.
        update_related_parts():
            Updates related parts to the Manifold in the database.
        update_manifold_assembly():
            Updates the assembly data for the Manifold in the database.
            Is in this class as the FR matching procedure is the final step
            before initiating the Manifold assembly.
    """

    def __init__(self, session: "Session", fms: "FMSDataStructure", fr_certifications: str = "certifications", anode_target_15: float = 3.006, anode_target_24: float = 4.809,
                 cathode_target_15: float = 0.231, cathode_target_24: float = 0.370, min_radius: float = 0.22, max_radius: float = 0.25):
        
        self.Session = session
        self.fms = fms
        self.fr_certifications = fr_certifications
        self.min_radius = min_radius
        self.max_radius = max_radius
        self.anode_target_15 = anode_target_15
        self.anode_target_24 = anode_target_24
        self.cathode_target_15 = cathode_target_15
        self.cathode_target_24 = cathode_target_24
        self.fr_id = None
        self.fr_test_results = {}
        self.fr_matching_dict = {}
        self.fr_matching_ratios = {}
        self.lines = ''
        self.manifold_assembly_data = {}
        self.xenon_density = 5.894  # kg/m^3 at 20°C and 1 atm
        self.krypton_density = 3.749  # kg/m^3 at 20°C and 1 atm

    def listen_to_fr_certifications(self) -> None:
        """
        Start listening for new FR certification files.
        """
        data_folder = os.path.join(os.getcwd(), self.fr_certifications)
        try:
            self.fr_listener = FRListener(data_folder)
            print(f"Started monitoring FR certifications in: {data_folder}")
            while True:
                try:
                    time.sleep(1)  # Keep the script running to monitor for new files
                    
                    # Check if listener has processed new data
                    if hasattr(self.fr_listener, 'processed') and self.fr_listener.processed:

                        if hasattr(self.fr_listener, 'fr_data') and self.fr_listener.fr_data:
                            self.certification = self.fr_listener.fr_data.certification
                            self.fr_info = self.fr_listener.fr_data.extracted_fr_parts

                            if not self.fr_listener.parts:
                                self.update_fr_certification()
                            else:
                                self.update_fr_part_certification()
                                self.fr_listener.parts = False
                            self.fr_listener.processed = False

                except Exception as e:
                    print(f"Error in fr listener loop: {str(e)}")
                    print("Listener will continue monitoring...")
                    traceback.print_exc()
                    
        except KeyboardInterrupt:
            print("Stopping fr test results listener...")
            if hasattr(self, 'fr_listener') and self.fr_listener:
                self.fr_listener.observer.stop()
                self.fr_listener.observer.join()
        except Exception as e:
            print(f"Fatal error in fr test results listener: {str(e)}")
            traceback.print_exc()
            # Try to restart the listener after a brief delay
            time.sleep(5)
            print("Attempting to restart fr test results listener...")
            self.listen_to_fr_certifications()

    def update_fr_certification(self, fr_data: FRData = None) -> None:
        """
        Update the database with new FR certification data.
        Args:
            fr_data (FRData, optional): FRData object holding the relevant processed data. Defaults to None.
            If fr_data is None, uses self.fr_data from the listener.
        """
        session = None
        self.fr_data = fr_data
        try:
            session: "Session" = self.Session()

            fr_parts = self.fr_data.extracted_fr_parts
            certification = self.fr_data.certification

            # ---- Original anode/cathode FR logic ----
            for fr_id, fr_item in fr_parts.items():
                part_type = fr_item.get("fr_type", "unknown")

                if part_type == "anode":
                    fr_model = AnodeFR(
                        fr_id=fr_id,
                        thickness=fr_item["thickness"],
                        orifice_diameter=fr_item["orifice"],
                        deviation=fr_item["deviation"],
                        status_geometry=fr_item["status_geometry"]
                    )
                    fr_cert = FRCertification(
                        certification=certification,
                        drawing=fr_item['drawing'],
                        part_name=FRParts.ANODE.value,
                        anode_fr_id=fr_id
                    )

                elif part_type == "cathode":
                    fr_model = CathodeFR(
                        fr_id=fr_id,
                        thickness=fr_item["thickness"],
                        orifice_diameter=fr_item["orifice"],
                        deviation=fr_item["deviation"],
                        status_geometry=fr_item["status_geometry"]
                    )
                    fr_cert = FRCertification(
                        certification=certification,
                        drawing=fr_item['drawing'],
                        part_name=FRParts.CATHODE.value,
                        cathode_fr_id=fr_id
                    )

                else:
                    fr_model = None
                    fr_cert = None

                if fr_model:
                    session.merge(fr_model)
                if fr_cert:
                    session.merge(fr_cert)

            # ---- Integrated generic FR part certification logic ----
            for part_name, part_data in fr_parts.items():
                drawing = part_data.get('drawing', 'unknown')
                amount = part_data.get('amount', 0)
                for _ in range(amount):
                    fr_cert = FRCertification(
                        certification=certification if certification else part_data.get("certification"),
                        drawing=drawing,
                        part_name=part_name
                    )
                    session.merge(fr_cert)

            # ---- Commit once at the end ----
            session.commit()

            # ---- Print tables ----
            self.fms.print_table(AnodeFR)
            self.fms.print_table(CathodeFR)
            self.fms.print_table(FRCertification)

        except Exception as e:
            print(f"Error updating FR certification: {e}")
            if session:
                session.rollback()
        finally:
            if session:
                session.close()

    def get_status(self, flow_rates: list, type: str) -> FRStatus:
        """
        Determines the status of FRs based on flow rates and type.
        Args:
            flow_rates (list): List of flow rates at different pressures.
            type (str): Type of FR ('Anode' or 'Cathode').
        Returns:
            FRStatus: Status of the FR based on flow rate differences.
        """
        if type == 'Anode':
            target_15 = self.anode_target_15
            target_24 = self.anode_target_24
        else:
            target_15 = self.cathode_target_15
            target_24 = self.cathode_target_24
        
        if len(flow_rates) == 4:
            diff_15 = abs(flow_rates[1] - target_15) / target_15 * 100
            diff_24 = abs(flow_rates[3] - target_24) / target_24 * 100

            avg_diff = (diff_15 + diff_24) / 2
            if avg_diff <= 15:
                return FRStatus.OK
            elif avg_diff > 15:
                return FRStatus.DIFF_FLOWRATE
        else:
            return None
        
    def get_radius_status(self, status: FRStatus, radius: float) -> FRStatus:
        """
        Adjusts the FR status based on the radius size and the current status.
        Args:
            status (FRStatus): Current status of the FR.
            radius (float): Measured radius of the FR. 
        Returns:
            FRStatus: Updated status considering the radius.
        """
        radius_different = radius < self.min_radius or radius > self.max_radius if radius else False

        if radius_different and not status == FRStatus.OK and not status == FRStatus.DIFF_GEOMETRY:
            return FRStatus.DIFF_GEOMETRY
        elif radius_different and status == FRStatus.OK:
            return FRStatus.DIFF_RADIUS
        elif not radius_different:
            return status

    def hagen_poiseuille(self, gas_type: str, flow_rate: list, thickness: float, orifice_diameter: float, viscosity: float = 1e-6) -> list:
        if any(val is None for val in flow_rate) or thickness is None or orifice_diameter is None or len(flow_rate) < 4:
            return None
        pressure_drop = viscosity*flow_rate*thickness/orifice_diameter**4*1000/self.xenon_density if gas_type == "Xe"\
              else viscosity*flow_rate*thickness/orifice_diameter**4*1000/self.krypton_density if gas_type == "Kr" else None
        return list(pressure_drop) if len(pressure_drop) >= 4 and thickness and orifice_diameter and len(flow_rate) >= 4 else None

    def fr_test_result_sql(self, test_results: dict, excel_extraction: bool, session: "Session") -> AnodeFR | CathodeFR | None:
        """
        Handles database insertion of FR test results.
        """
        if not test_results:
            return None

        fr_type = test_results['fr']
        model_cls = AnodeFR if fr_type == "Anode" else CathodeFR
        existing_entry = session.query(model_cls).filter_by(fr_id=test_results['serial_number']).first()

        thickness = existing_entry.thickness if existing_entry else test_results.get('thickness', None)
        orifice = existing_entry.orifice_diameter if existing_entry else test_results.get('orifice_diameter', None)

        fr_model = model_cls(
            fr_id=test_results['serial_number'],
            pressures=test_results['pressures'],
            flow_rates=test_results['flow_rates'],
            temperature=test_results['temperature'],
            drawing=test_results['drawing'],
            date=datetime.fromisoformat(test_results['date']).date() if test_results['date'] else None,
            remark=test_results['remark'],
            pressure_drop=self.hagen_poiseuille(test_results.get('gas_type', "Xe"), \
                                                np.array(test_results['flow_rates']), thickness, orifice) if thickness and orifice else None,
            status=self.get_status(test_results['flow_rates'], fr_type),
            gas_type=test_results.get('gas_type', "Xe"),
            allocated=test_results.get('allocated'),
            set_id=test_results.get('set_id'),
            trs_reference=test_results.get('trs_reference', ""),
            tools=test_results.get('tools', []),
            orifice_diameter=orifice,
            radius=test_results.get('radius', None),
            thickness=thickness,
            deviation=test_results.get('deviation', None),
            operator=test_results.get('operator', self.fms.operator)
        )

        if "radius" in test_results and "status_geometry" in test_results:
            fr_model.status_geometry = self.get_radius_status(test_results['status_geometry'], test_results['radius'])

        return fr_model
    
    def update_fr_test_tools(self, tools: list[dict[str, str]]) -> None:
        """
        Updates the tools used during an FR test in the testing_tools table.
        
        :param tools: List of the tools used during the FR test.
        :type tools: list[dict[str, str]]
        """
        session = None
        try:
            session: "Session" = self.Session()
            for tool in tools:
                description = tool.get("description")
                model = tool.get("model")
                serial = tool.get("serial_number")
                range_ = tool.get("equipment_range")
                accuracy = tool.get("accuracy")
                last_calibration_date = tool.get("last_calibration_date")
                next_calibration_date = tool.get("next_calibration_date")
                
                existing_tool = session.query(TestingTools).filter_by(description=description, model=model, serial_number=serial).first()
                if existing_tool:
                    existing_tool.equipment_range = range_
                    existing_tool.accuracy = accuracy
                    existing_tool.last_calibration_date = last_calibration_date
                    existing_tool.next_calibration_date = next_calibration_date

                else:
                    new_entry = TestingTools(
                        description = description,
                        model = model,
                        serial_number = serial,
                        equipment_range = range_,
                        accuracy = accuracy,
                        last_calibration_date = last_calibration_date,
                        next_calibration_date = next_calibration_date
                    )
                    session.add(new_entry)
            session.commit()
        except Exception as e:
            print(f"Error updating FR test results: {e}")
            if session:
                session.rollback()
            traceback.print_exc()
        finally:
            if session:
                session.close()

    def update_fr_test_results(self, excel_extraction: bool = False, anode_path: str = None,\
                                cathode_path: str = None, operator: str = "JKR", tools_path: str = "") -> None:
        """
        Update the FR test results in the database.
        """
        session = None
        try:
            session: "Session" = self.Session()

            if excel_extraction:
                self.fr_data = FRData(anode_excel=anode_path, cathode_excel=cathode_path) if anode_path and cathode_path else FRData()
                self.fr_test_results = self.fr_data.extract_data_from_excel(operator = operator, tools_path=tools_path)
                if not self.fr_test_results:
                    print("No FR test results to update.")
                    return
                for test_result in self.fr_test_results:
                    fr_model = self.fr_test_result_sql(test_result, excel_extraction, session)
                    if fr_model:
                        session.merge(fr_model)
            else:
                if not self.fr_test_results:
                    print("No FR test results to update.")
                    return
                fr_model = self.fr_test_result_sql(self.fr_test_results, excel_extraction, session)
                if fr_model:
                    session.merge(fr_model)

            session.commit()

            # self.fms.print_table(AnodeFR)
            # self.fms.print_table(CathodeFR)
        except Exception as e:
            print(f"Error updating FR test results: {e}")
            if session:
                session.rollback()
            traceback.print_exc()
        finally:
            if session:
                session.close()

    def update_related_parts(self, session: "Session", set_id: int, anode_fr: int, anode_filter: str, \
                             anode_outlet: str, cathode_fr: int, cathode_filter: str, cathode_outlet: str, lpt_id: int) -> None:
        
        """
        Update related parts to the Manifold in the database.
        Args:
            session (Session): SQLAlchemy session object.
            set_id (int): Set ID of the Manifold.
            anode_fr (int): Serial number of the anode flow restrictor.
            anode_filter (str): Certification of the anode filter.
            anode_outlet (str): Certification of the anode outlet.
            cathode_fr (int): Serial number of the cathode flow restrictor.
            cathode_filter (str): Certification of the cathode filter.
            cathode_outlet (str): Certification of the cathode outlet.
            lpt_id (int): Serial number of the LPT calibration.
        """

        anode = session.query(AnodeFR).filter_by(fr_id=anode_fr).first()
        if anode:
            anode.set_id = set_id
        else:
            print(f"Anode FR with ID {anode_fr} not found in database, likely has not been flow tested.")

        cathode = session.query(CathodeFR).filter_by(fr_id=cathode_fr).first()
        if cathode:
            cathode.set_id = set_id
        else:
            print(f"Cathode FR with ID {cathode_fr} not found in database, likely has not been flow tested.")


        # Get available entries for filter and outlet
        filter_entries = session.query(FRCertification).filter_by(
            certification=anode_filter,
            part_name='ejay filter',
            anode_fr_id=None,
            cathode_fr_id=None
        ).limit(2).all()

        outlet_entries = session.query(FRCertification).filter_by(
            certification=anode_outlet,
            part_name='restrictor outlet',
            anode_fr_id=None,
            cathode_fr_id=None
        ).limit(2).all()

        if filter_entries:
            anode_entry = filter_entries[0]
            anode_entry.anode_fr_id = anode_fr

            if len(filter_entries) > 1:
                cathode_entry = filter_entries[1]
                cathode_entry.cathode_fr_id = cathode_fr
            else:
                print(f"Cathode filter with certification {cathode_filter} not found in database")
        else:
            print(f"Anode filter with certification {anode_filter} not found in database")
            print(f"Cathode filter with certification {cathode_filter} not found in database")

        if outlet_entries:
            anode_outlet_entry = outlet_entries[0]
            anode_outlet_entry.anode_fr_id = anode_fr

            if len(outlet_entries) > 1:
                cathode_outlet_entry = outlet_entries[1]
                cathode_outlet_entry.cathode_fr_id = cathode_fr
            else:
                print(f"Cathode outlet with certification {cathode_outlet} not found in database")
        else:
            print(f"Anode outlet with certification {anode_outlet} not found in database")
            print(f"Cathode outlet with certification {cathode_outlet} not found in database")

        if lpt_id:
            lpt_calibration = session.query(LPTCalibration).filter_by(set_id=None, lpt_id=lpt_id).first()
            if lpt_calibration:
                lpt_calibration.set_id = set_id

    def update_manifold_assembly(self) -> None:
        """
        Update the manifold assembly data in the database.
        """
        session: "Session" = None
        try:
            session = self.Session()
            set_id = self.manifold_assembly_data['set_id']
            manifold_certification = self.manifold_assembly_data['manifold_certification']
            anode_fr = self.manifold_assembly_data['anode_fr']
            anode_filter = self.manifold_assembly_data['anode_filter']
            anode_outlet = self.manifold_assembly_data['anode_outlet']
            cathode_fr = self.manifold_assembly_data['cathode_fr']
            cathode_filter = self.manifold_assembly_data['cathode_filter']
            cathode_outlet = self.manifold_assembly_data['cathode_outlet']
            lpt_id = self.manifold_assembly_data['lpt_id']
            ac_ratio_specified = self.manifold_assembly_data.get('ac_ratio_specified', 13)
            ac_ratio = self.manifold_assembly_data.get('ac_ratio', 13)

            available_entry = session.query(ManifoldStatus).filter_by(set_id=None, allocated=None,\
                                                                       status=ManifoldProgressStatus.AVAILABLE, certification=manifold_certification).first() 
            if available_entry:
                available_entry.set_id = set_id
                available_entry.ac_ratio = ac_ratio
                available_entry.ac_ratio_specified = ac_ratio_specified
                available_entry.status = ManifoldProgressStatus.ASSEMBLY_COMPLETED
            else:
                new_entry = ManifoldStatus(
                    set_id=set_id,
                    certification=manifold_certification,
                    status=ManifoldProgressStatus.ASSEMBLY_COMPLETED,
                    ac_ratio=ac_ratio,
                    ac_ratio_specified=ac_ratio_specified,
                )
                session.add(new_entry)

            self.update_related_parts(
                    session,
                    set_id,
                    anode_fr,
                    anode_filter,
                    anode_outlet,
                    cathode_fr,
                    cathode_filter,
                    cathode_outlet,
                    lpt_id
                )
                    
            session.commit()

        except Exception as e:
            print(f"Error updating manifold assembly data: {str(e)}")
            if session:
                session.rollback()
            traceback.print_exc()
        finally:
            if session:
                session.close()

if __name__ == "__main__":
    # file = "certifications\C25-1033 Veldlaser 514422.pdf"
    # file = "certifications\C25-1034 Veldlaser 514422.pdf"
    file = "certifications\C25-0333 Veldlaser 513226.pdf"

    fr_data = FRData(pdf_file = file)
    company = "Veldlaser"
    reader = TextractReader(pdf_file=file, bucket_folder="Certifications", company=company)
    total_lines = reader.get_text()
    fr_data.get_certification(total_lines)
    # fr_data.get_ocr_certification()
    print(fr_data.extracted_fr_parts)
